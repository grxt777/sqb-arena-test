"""
Парсер XLSX с реестром банкоматов.

Ожидаемая структура (нумерация по EXCEL-столбцам, 1-based):
  A — Номер банкомата (или просто "№")
  B — Номер по филиалам (или просто "№")
  C — Локал код
  D — Худуд / Область / Регион
  E — Филиал
  F — Модели (модель банкомата)
  G — Тулов тизими тури (тип сети)
  H — Серия раками
  I — Merchant id
  J — Term id (терминал ID)
  K — Манзили (адрес)
  L — Геолокация (alt/lat)
  M — Геолокация (long/lon)

Первая строка — заголовок.
"""

from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.utils import get_column_letter

from .config import DEFAULT_CAPACITY

log = logging.getLogger(__name__)


# Синонимы заголовков (в lower-case, без пробелов) → каноническое имя поля
HEADER_ALIASES: Dict[str, str] = {
    # ── ATM number (колонка A) ──────────────────────────────
    "номербанкомата": "atm_number",
    "номер_банкомата": "atm_number",
    "atm_number": "atm_number",
    "atmnumber": "atm_number",
    "atm": "atm_number",
    "номератм": "atm_number",
    "номер": "atm_number",
    "nomer": "atm_number",   # нормализованный "№"
    "no": "atm_number",
    "": "atm_number",   # на случай пустой шапки
    "n": "atm_number",

    # ── Branch code (колонка B) ────────────────────────────
    "номерпофилиалам": "branch_code",
    "номер_по_филиалам": "branch_code",
    "branch_code": "branch_code",
    "branchcode": "branch_code",
    "филиалкод": "branch_code",
    "филиал_код": "branch_code",
    "кодфилиала": "branch_code",

    # ── Local code (колонка C) ─────────────────────────────
    "локалкод": "local_code",
    "локал_код": "local_code",
    "local_code": "local_code",
    "localcode": "local_code",
    "локал": "local_code",

    # ── Region (колонка D) — Худуд / Область / Вилоят ──────
    "худуд": "region",
    "область": "region",
    "region": "region",
    "viloyat": "region",
    "вилоят": "region",
    "район": "region",
    "district": "region",

    # ── Branch (колонка E) ─────────────────────────────────
    "филиал": "branch",
    "branch": "branch",
    "filial": "branch",

    # ── Model (колонка F) — Модель / Модели ────────────────
    "модельбанкомата": "model",
    "модель_банкомата": "model",
    "модели": "model",
    "модел": "model",
    "model": "model",
    "модель": "model",

    # ── Network type (колонка G) — Тулов / Тармоқ ─────────
    "туловтизимитури": "network_type",
    "тулов_тизими_тури": "network_type",
    "тулов": "network_type",
    "тармоқтизимитури": "network_type",
    "тармоктизимитури": "network_type",
    "network_type": "network_type",
    "networktype": "network_type",
    "тармок": "network_type",

    # ── Serial (колонка H) ─────────────────────────────────
    "серияраками": "serial",
    "серия_раками": "serial",
    "серия_рақами": "serial",
    "serial": "serial",
    "серия": "serial",
    "sn": "serial",

    # ── Merchant ID (колонка I) ───────────────────────────
    "мерчантайди": "merchant_id",
    "мерчант_айди": "merchant_id",
    "мерчантid": "merchant_id",
    "merchant_id": "merchant_id",
    "merchantid": "merchant_id",
    "merchant": "merchant_id",
    "mid": "merchant_id",

    "терминалайди": "terminal_id",
    "терминал_айди": "terminal_id",
    "терминалid": "terminal_id",
    "terminal_id": "terminal_id",
    "terminalid": "terminal_id",
    "term_id": "terminal_id",
    "termid": "terminal_id",
    "tеrmid": "terminal_id",
    "tеrm_id": "terminal_id",
    "tеrm": "terminal_id",
    "терминал": "terminal_id",
    "tid": "terminal_id",
    "term": "terminal_id",

    # ── Address (колонка K) — Манзили / Адрес ─────────────
    "манзили": "address",
    "манзил": "address",
    "адрес": "address",
    "address": "address",
    "manzil": "address",

    # ── Lat / Lon (колонки L, M) — Геолокация (alt/long) ──
    # alt = "altitude" (сокращение), в нашей таблице = latitude
    "геолокация(alt)": "lat",
    "геолокация_alt": "lat",
    "геолокация(лат)": "lat",
    "геолокация(широта)": "lat",
    "геолокацияalt": "lat",
    "геолокацияlat": "lat",
    "геолокация": "lat",
    "alt": "lat",
    "latitude": "lat",
    "lat": "lat",
    "широта": "lat",

    "геолокация(long)": "lon",
    "геолокация_long": "lon",
    "геолокация(долгота)": "lon",
    "геолокацияlong": "lon",
    "геолокацияlon": "lon",
    "long": "lon",
    "longitude": "lon",
    "lon": "lon",
    "lng": "lon",
    "долгота": "lon",
}


REQUIRED_FIELDS = ("terminal_id",)


def _normalize_header(text: Any) -> str:
    """Нормализует заголовок: нижний регистр, без пробелов/знаков препинания."""
    if text is None:
        return ""
    s = str(text).strip().lower()
    # Схлопываем кириллические спец-символы к их ASCII-аналогам
    cyr_map = {
        "қ": "к", "қ": "к", "ҳ": "х", "ғ": "г",
        "ў": "у", "қ": "к", "Қ": "к", "Ҳ": "х",
        "Ў": "у", "Ғ": "г",
    }
    for src, dst in cyr_map.items():
        s = s.replace(src, dst)
    # Спец-символ "номер по порядку" → канонический маркер "nomer"
    s = s.replace("№", "nomer")
    s = re.sub(r"[\s_\-\.,()]+", "", s)
    s = s.replace("'", "").replace("`", "").replace("'", "")
    return s


def _parse_coord(value: Any) -> Optional[float]:
    """Парсит координату из строки или числа."""
    if value is None or value == "":
        return None
    try:
        v = float(value)
    except (TypeError, ValueError):
        if isinstance(value, str):
            cleaned = re.sub(r"[^\d.\-]", "", value.replace(",", "."))
            try:
                v = float(cleaned)
            except (TypeError, ValueError):
                return None
        else:
            return None
    if -180 <= v <= 180:
        return round(v, 6)
    return None


def _parse_int(value: Any) -> Optional[int]:
    if value is None or value == "":
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        if isinstance(value, str):
            cleaned = re.sub(r"[^\d\-]", "", value)
            if cleaned.lstrip("-").isdigit():
                try:
                    return int(cleaned)
                except ValueError:
                    return None
        return None


def _parse_str(value: Any) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def _match_headers(ws) -> Tuple[Dict[int, str], int]:
    """
    Возвращает:
      - словарь {column_index_1based: canonical_field_name}
      - номер строки заголовка (1-based)

    Если в первой строке не нашлось ни одного знакомого заголовка —
    считаем, что первая строка — данные, и поля выводятся по позиции.

    Особый случай: две колонки с заголовком "№" (atm_number и branch_code).
    Первой "№" → atm_number, второй "№" → branch_code.
    """
    header_row_idx = None
    for row_idx in range(1, min(6, ws.max_row + 1)):
        row = next(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))
        normalized = [_normalize_header(c) for c in row]
        hits = sum(1 for n in normalized if n in HEADER_ALIASES)
        if hits >= 3:
            header_row_idx = row_idx
            break

    column_map: Dict[int, str] = {}

    if header_row_idx is None:
        # fallback: маппим по позициям A..M
        position_map = {
            1: "atm_number",
            2: "branch_code",
            3: "local_code",
            4: "region",
            5: "branch",
            6: "model",
            7: "network_type",
            8: "serial",
            9: "merchant_id",
            10: "terminal_id",
            11: "address",
            12: "lat",
            13: "lon",
        }
        for col_idx, field in position_map.items():
            column_map[col_idx] = field
        log.warning(
            "Заголовки XLSX не распознаны — используется позиционный маппинг A..M"
        )
        return column_map, 1

    header_row = next(
        ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True)
    )

    # Сначала разбираем колонки, у которых есть уникальные алиасы
    for col_idx, cell in enumerate(header_row, start=1):
        norm = _normalize_header(cell)
        if norm in HEADER_ALIASES:
            field = HEADER_ALIASES[norm]
            if field not in column_map.values():
                column_map[col_idx] = field
            # дубликаты atm_number → branch_code
            elif field == "atm_number" and "branch_code" not in column_map.values():
                column_map[col_idx] = "branch_code"
            # остальные дубли — игнорируем

    # Эвристика: если в левых 2-3 колонках все значения нормализуются
    # в "номер" / "no" / "n" (т.е. две "№") и ни atm_number, ни branch_code
    # ещё не сопоставлены — ставим их по позиции
    if "atm_number" not in column_map.values():
        for col_idx, cell in enumerate(header_row, start=1):
            norm = _normalize_header(cell)
            if norm in ("", "n", "no", "номер"):
                column_map[col_idx] = "atm_number"
                break
    if "branch_code" not in column_map.values():
        for col_idx, cell in enumerate(header_row, start=1):
            norm = _normalize_header(cell)
            if norm in ("", "n", "no", "номер"):
                column_map[col_idx] = "branch_code"
                break

    return column_map, header_row_idx


def parse_xlsx(
    path: str | Path,
    sheet_name: Optional[str] = None,
) -> Dict[str, Any]:
    """
    Парсит XLSX и возвращает:
      {
        "records":     List[Dict]   — распарсенные записи
        "errors":      List[Dict]   — список ошибок построчно
        "total_rows":  int
        "header_row":  int
        "columns":     Dict[int, str]   — маппинг колонок
      }
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Файл не найден: {path}")

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    column_map, header_row = _match_headers(ws)

    if "terminal_id" not in column_map.values():
        header_vals = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
        detected = []
        for col_idx, cell in enumerate(header_vals, start=1):
            norm = _normalize_header(cell)
            detected.append(f"{get_column_letter(col_idx)}: '{cell}' -> '{norm}'")
        wb.close()
        raise ValueError(
            "Не удалось определить колонку Terminal ID. "
            f"Распознанная строка заголовка #{header_row}: [" + ", ".join(detected) + "]"
        )


    records: List[Dict[str, Any]] = []
    errors: List[Dict[str, Any]] = []
    seen_terminal_ids: set[str] = set()
    total_rows = 0

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1
    ):
        total_rows += 1
        if not row or all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
            continue

        record: Dict[str, Any] = {}
        for col_idx, field in column_map.items():
            if col_idx - 1 >= len(row):
                value = None
            else:
                value = row[col_idx - 1]
            if field in ("lat", "lon"):
                record[field] = _parse_coord(value)
            elif field == "capacity":
                record[field] = _parse_int(value) or DEFAULT_CAPACITY
            else:
                record[field] = _parse_str(value)

        terminal_id = record.get("terminal_id")
        if not terminal_id:
            errors.append({
                "row": row_idx,
                "error": "Пустой Terminal ID",
            })
            continue

        if terminal_id in seen_terminal_ids:
            errors.append({
                "row": row_idx,
                "terminal_id": terminal_id,
                "error": "Дубликат Terminal ID в файле",
            })
            continue
        seen_terminal_ids.add(terminal_id)

        record.setdefault("capacity", DEFAULT_CAPACITY)
        records.append(record)

    wb.close()

    return {
        "records": records,
        "errors": errors,
        "total_rows": total_rows,
        "header_row": header_row,
        "columns": {get_column_letter(k): v for k, v in column_map.items()},
    }


# ═══════════════════════════════════════════════════════════════════
# ИМПОРТ ФИЛИАЛОВ (branches)
# ═══════════════════════════════════════════════════════════════════
#
# Ожидаемая структура XLSX:
#   № — порядковый номер
#   Локал код — уникальный код филиала
#   Регион — область / регион
#   Адрес — адрес филиала
#   Lat — широта
#   Lon — долгота
#   Инкассация — 0 (не предназначен для выезда инкассаторов) / 1 (предназначен)

BRANCH_HEADER_ALIASES: Dict[str, str] = {
    # № — порядковый номер
    "nomer": "number",
    "no": "number",
    "n": "number",
    "номер": "number",
    "": "number",

    # Локал код
    "локалкод": "local_code",
    "локал_код": "local_code",
    "local_code": "local_code",
    "localcode": "local_code",
    "локал": "local_code",

    # Регион
    "регион": "region",
    "region": "region",
    "область": "region",
    "худуд": "region",
    "viloyat": "region",
    "вилоят": "region",

    # Адрес
    "адрес": "address",
    "address": "address",
    "манзили": "address",
    "манзил": "address",
    "manzil": "address",

    # Lat
    "lat": "lat",
    "latitude": "lat",
    "широта": "lat",

    # Lon
    "lon": "lon",
    "long": "lon",
    "longitude": "lon",
    "lng": "lon",
    "долгота": "lon",

    # Инкассация
    "инкассация": "incassation",
    "incassation": "incassation",
    "инкассацио": "incassation",
}

BRANCH_REQUIRED_FIELDS = ("local_code",)


def _match_branch_headers(ws) -> Tuple[Dict[int, str], int]:
    """
    Аналог _match_headers, но для реестра филиалов.
    Ожидаемые колонки: № локал код регион адрес lat lon инкассация
    """
    header_row_idx = None
    for row_idx in range(1, min(6, ws.max_row + 1)):
        row = next(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))
        normalized = [_normalize_header(c) for c in row]
        hits = sum(1 for n in normalized if n in BRANCH_HEADER_ALIASES)
        if hits >= 3:
            header_row_idx = row_idx
            break

    column_map: Dict[int, str] = {}

    if header_row_idx is None:
        # fallback: позиционный маппинг № локал_код регион адрес lat lon инкассация
        position_map = {
            1: "number",
            2: "local_code",
            3: "region",
            4: "address",
            5: "lat",
            6: "lon",
            7: "incassation",
        }
        for col_idx, field in position_map.items():
            column_map[col_idx] = field
        log.warning(
            "Заголовки XLSX (филиалы) не распознаны — используется позиционный маппинг"
        )
        return column_map, 1

    header_row = next(
        ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True)
    )

    for col_idx, cell in enumerate(header_row, start=1):
        norm = _normalize_header(cell)
        if norm in BRANCH_HEADER_ALIASES:
            field = BRANCH_HEADER_ALIASES[norm]
            if field not in column_map.values():
                column_map[col_idx] = field

    if "number" not in column_map.values():
        for col_idx, cell in enumerate(header_row, start=1):
            norm = _normalize_header(cell)
            if norm in ("", "n", "no", "номер"):
                column_map[col_idx] = "number"
                break

    return column_map, header_row_idx


def _parse_incassation(value: Any) -> int:
    """Парсит колонку 'Инкассация': 0 — не выездной, 1 — выездной."""
    if value is None or value == "":
        return 0
    s = str(value).strip().lower()
    if s in ("1", "1.0", "true", "да", "yes", "ha", "ha'"):
        return 1
    if s in ("0", "0.0", "false", "нет", "no", "yoq", "йук"):
        return 0
    try:
        return 1 if int(float(value)) != 0 else 0
    except (TypeError, ValueError):
        return 0


def parse_branches_xlsx(
    path: str | Path,
    sheet_name: Optional[str] = None,
) -> Dict[str, Any]:
    """
    Парсит XLSX с реестром филиалов и возвращает:
      {
        "records":     List[Dict]   — распарсенные записи
        "errors":      List[Dict]   — список ошибок построчно
        "total_rows":  int
        "header_row":  int
        "columns":     Dict[int, str]   — маппинг колонок
      }

    Колонки: № локал код регион адрес lat lon инкассация
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Файл не найден: {path}")

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    column_map, header_row = _match_branch_headers(ws)

    if "local_code" not in column_map.values():
        header_vals = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
        detected = []
        for col_idx, cell in enumerate(header_vals, start=1):
            norm = _normalize_header(cell)
            detected.append(f"{get_column_letter(col_idx)}: '{cell}' -> '{norm}'")
        wb.close()
        raise ValueError(
            "Не удалось определить колонку 'Локал код'. "
            f"Распознанная строка заголовка #{header_row}: [" + ", ".join(detected) + "]"
        )

    records: List[Dict[str, Any]] = []
    errors: List[Dict[str, Any]] = []
    seen_local_codes: set[str] = set()
    total_rows = 0

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1
    ):
        total_rows += 1
        if not row or all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
            continue

        record: Dict[str, Any] = {}
        for col_idx, field in column_map.items():
            if col_idx - 1 >= len(row):
                value = None
            else:
                value = row[col_idx - 1]
            if field in ("lat", "lon"):
                record[field] = _parse_coord(value)
            elif field == "incassation":
                record[field] = _parse_incassation(value)
            else:
                record[field] = _parse_str(value)

        local_code = record.get("local_code")
        if not local_code:
            errors.append({
                "row": row_idx,
                "error": "Пустой Локал код",
            })
            continue

        if local_code in seen_local_codes:
            errors.append({
                "row": row_idx,
                "local_code": local_code,
                "error": "Дубликат Локал кода в файле",
            })
            continue
        seen_local_codes.add(local_code)

        record.setdefault("incassation", 0)
        records.append(record)

    wb.close()

    return {
        "records": records,
        "errors": errors,
        "total_rows": total_rows,
        "header_row": header_row,
        "columns": {get_column_letter(k): v for k, v in column_map.items()},
    }
